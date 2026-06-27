"""
generate_xlsx.py

Lee el store consolidado (datos_consolidados/todos_registros.json) y genera
un archivo Excel con todos los registros de todas las fuentes.

Requiere: pip install openpyxl
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl no está instalado. Ejecuta: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

BASE_DIR = Path(__file__).parent
INPUT_FILE = BASE_DIR / "datos_consolidados" / "todos_registros.json"
OUTPUT_FILE = BASE_DIR / "datos_consolidados" / "todos_registros.xlsx"

COLUMNS = [
    "id", "nombre", "cedula", "edad", "ultima_ubicacion",
    "telefono_contacto", "observaciones", "estado",
    "ubicacion_encontrado", "encontrado_por", "encontrado_por_cedula",
    "foto_url", "fecha_registro", "fecha_actualizacion", "es_menor", "fuente"
]

COLUMN_LABELS = {
    "id": "ID",
    "nombre": "Nombre",
    "cedula": "Cédula",
    "edad": "Edad",
    "ultima_ubicacion": "Última Ubicación",
    "telefono_contacto": "Teléfono Contacto",
    "observaciones": "Observaciones",
    "estado": "Estado",
    "ubicacion_encontrado": "Ubicación Encontrado",
    "encontrado_por": "Encontrado Por",
    "encontrado_por_cedula": "Cédula Encontrado",
    "foto_url": "URL Foto",
    "fecha_registro": "Fecha Registro",
    "fecha_actualizacion": "Fecha Actualización",
    "es_menor": "Es Menor",
    "fuente": "Fuente"
}

COL_WIDTHS = {
    "id": 20, "nombre": 30, "cedula": 15, "edad": 8,
    "ultima_ubicacion": 30, "telefono_contacto": 18,
    "observaciones": 45, "estado": 14,
    "ubicacion_encontrado": 25, "encontrado_por": 22,
    "encontrado_por_cedula": 18, "foto_url": 35,
    "fecha_registro": 24, "fecha_actualizacion": 24,
    "es_menor": 10, "fuente": 24
}


def main():
    print("====================================================")
    print("           GENERADOR DE ARCHIVO XLSX                ")
    print("====================================================")

    if not INPUT_FILE.exists():
        print(f"Error: No se encontró {INPUT_FILE}", file=sys.stderr)
        print("Asegúrate de que run_daily.py haya ejecutado el paso de consolidación.", file=sys.stderr)
        sys.exit(1)

    print(f"Cargando: {INPUT_FILE}")
    data = json.loads(INPUT_FILE.read_text(encoding="utf-8"))
    print(f"Total de registros: {len(data)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personas"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_key in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=COLUMN_LABELS.get(col_key, col_key))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_key, 15)

    ws.row_dimensions[1].height = 30

    for row_idx, record in enumerate(data, start=2):
        for col_idx, col_key in enumerate(COLUMNS, start=1):
            val = record.get(col_key)
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            ws.cell(row=row_idx, column=col_idx, value=val)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_FILE)
    print(f"XLSX generado: {OUTPUT_FILE}")
    print(f"Filas de datos: {len(data)}")


if __name__ == "__main__":
    main()
